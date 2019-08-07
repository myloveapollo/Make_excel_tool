import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers
import datetime
from openpyxl.worksheet.datavalidation import DataValidation
pd.options.mode.chained_assignment = None


def cell_style(ws, len_index):
	width_dict = {'A': 155}
	font = Font(name='宋体', size=46, bold=True)
	# thin = Side(border_style='thin', color='00000000')
	alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
	for row in ws.iter_rows(min_row=1, max_row=len_index, min_col=1, max_col=1):
		for cell in row:
			cell.font = font
			# cell.border = Border(bottom=thin)
			cell.alignment = alignment
	for k, v in width_dict.items():
		ws.column_dimensions[k].width = v
	for i in range(len_index + 1):
		ws.row_dimensions[i].height = 409


def read_data(file_name, finish_name):
	data = pd.read_excel(file_name)  # usecols=[2,0,1,6,7,3,4]

	data['班次'] = data['班次'].str.replace("[(].*[)]", '', regex=True)
	data['教室'] = data['教室'].str.replace("[【].*[】]|[(].*[)]", '', regex=True)
	data['教师'] = data['教师'].str.replace("[0-9]*$", '', regex=True)
	data['辅导老师'] = data['辅导老师'].str.replace("[0-9]*$", '', regex=True)

	data_now= []
	data_add = []
	for n in data['辅导老师']:
		if n != ' ':
			a = '辅导老师:' + n
			data_add.append('主讲老师:')
		else:
			a = n
			data_add.append('老师:')
		data_now.append(a)

	make_data = data['班次'] + '\n' + data['年级'] + '   ' + data['学科'] + '\n' + data['教室'] + '\n' + data[
		'上课时间'] + '\n' + data_add + data['教师'] + '\n' + data_now
	make_data = pd.DataFrame(make_data)

	wb = Workbook()
	ws = wb.create_sheet('Sheet0', -1)

	for r in dataframe_to_rows(make_data, index=False, header=False):
		ws.append(r)
	cell_style(ws, len(make_data.index))
	wb.save(finish_name)


# finish_name = 'ceshi.xlsx'
# file_name = '新诊.xlsx'
# read_data(file_name)
