# -*-coding:GBK -*-
# ��Ӫ�����Űർ���
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers
from openpyxl.worksheet.datavalidation import DataValidation

pd.options.mode.chained_assignment = None


def cell_style(ws, len_index):
    width_dict = {'A': 15.13, 'B': 15.13, 'C': 19.28, 'D': 15.13, 'E': 15.13, 'F': 15.13}
    font = Font(name='����', size=11, bold=False)
    font_red = Font(color='FF0000')

    dv_type = DataValidation(type="list", formula1='"���,����,��Ϣ,����ְȱ��,��ѵ,����,ҽ����,�¼�,���,����,����,�����,ɥ��,�����"', allow_blank=False)
    ws.add_data_validation(dv_type)
    type_index = 'D2:D' + str(len_index + 1)
    dv_type.add(type_index)

    for row in ws.iter_rows(min_row=2, max_row=len_index + 1, min_col=1, max_col=1):  # ��һ�и�ʽ���Բ���6��8�ַ��ĸ�ʽ�ж�
        for cell in row:
            if len(cell.value) == 6 or len(cell.value) == 8:
                cell.font = font
            else:
                cell.font = font_red
            cell.number_format = '@'

    for row in ws.iter_rows(min_row=2, max_row=len_index + 1, min_col=2, max_col=3):  # 2/3�ı���ʽ
        for cell in row:
            cell.font = font
            cell.number_format = '@'  # 'yyyy-mm-dd'

    for row in ws.iter_rows(min_row=2, max_row=len_index + 1, min_col=4, max_col=4):  # 4�ı���ʽ
        for cell in row:
            cell.font = font

    for row in ws.iter_rows(min_row=2, max_row=len_index + 1, min_col=5, max_col=6):  # 5/6�и�ʽ
        for cell in row:
            cell.number_format = '@'
            if len(str(cell.value)) != 5:
                cell.font = font_red
            else:
                cell.font = font

    for k, v in width_dict.items():
        ws.column_dimensions[k].width = v


def wash_data(file_name_name):
    names = ['����', '����', '��һ', '�ܶ�', '����', '����', '����', '����', '����']

    # ��������
    data_name = pd.read_excel(file_name_name, sheet_name=0, header=0, names=None, usecols=[0])
    column = []
    for col in data_name:
        column.append(col)
    column_name = column[0].replace(' ', '').replace("��������", '').replace("�Ű��", '')
    in_excel = column_name + '���ڵ����.xlsx'  # ����������ɵı�  str(data_name.columns)

    data = pd.read_excel(file_name_name, sheet_name=0, header=None, names=names,
                         usecols=[1, 2, 3, 6, 9, 12, 15, 18, 21])  # ��ȡ��,
    data.replace('��', ':', regex=True, inplace=True)  # �����滻Ϊ:
    data_time = data.iloc[1]  # ��ȡ�����ѱ���,��ʽҪ������

    data.drop(axis=0, index=[1, 0, 2, 3, 4], inplace=True)  # ɾ��û�е�����Ϣ
    data.dropna(axis=0, how='all', inplace=True)  # ɾ��ȫ�յ�����Ϣ
    work_type = ['OFF', '���', '����ְȱ��',  '����', 'ҽ����', '�¼�', '���', '����', '����', '�����', 'ɥ��', '�����' ]
    for dub in work_type:
        data.replace(dub, dub+'-'+dub, regex=True, inplace=True)  # OFF����ڿ� ������ʽ�ӣ������ִ�Сд
    data.replace(np.nan, '�˴��Ű಻��Ϊ�յ�', regex=True, inplace=True)
    data = data.T  # ����

    # ���һ����
    data_make = pd.DataFrame(np.full([7 * len(data.columns), 6], np.nan),
                             columns=['����', '����', '����(YYYY-MM-DD)', '����', '�ϰ�ʱ��', '�°�ʱ��'])

    data_job_num = []
    for nam in list(data.iloc[0]):
        for i in range(1, 8):
            data_job_num.append(nam)
    data_make['����'] = data_job_num

    # ��ȡ����
    data_col_name = []
    for nam in list(data.iloc[1]):
        for i in range(1, 8):
            data_col_name.append(nam)
    data_make['����'] = pd.Series(data_col_name)  # ��������

    # ��ȡʱ��
    data.drop(axis=0, index='����', inplace=True)
    data.drop(axis=0, index='����', inplace=True)
    data_col_arrivetime = []  # �ϰ�ʱ��
    data_col_leavetime = []  # �°�ʱ��
    data_type = []  # ����

    for col in list(data.columns):
        data_split_f = data[col].str.split('-').str[0]  # �ϰ�ʱ����ǰ���
        data_split_s = data[col].str.split('-').str[1]  # �°�ʱ���ú����
        for f in list(data_split_f):
            if len(str(f)) == 4:  # len(f) == 4:
                f = '0' + f
            data_col_arrivetime.append(f)
        for s in list(data_split_s):
            data_col_leavetime.append(s)
            if s == 'OFF':
                data_type.append('��Ϣ')
            elif s =='���':
                data_type.append('���')
            elif s == '����ְȱ��':
                data_type.append('����ְȱ��')
            # elif s == '��ѵ':
            #     data_type.append('��ѵ')
            elif s == '����':
                data_type.append('����')
            elif s== 'ҽ����':
                data_type.append('ҽ����')
            elif s == '�¼�':
                data_type.append('�¼�')
            elif s == '���':
                data_type.append('���')
            elif s == '����':
                data_type.append('����')
            elif s == '����':
                data_type.append('����')
            elif s == '�����':
                data_type.append('�����')
            elif s == 'ɥ��':
                data_type.append('ɥ��')
            elif s == '�����':
                data_type.append('�����')
            else:
                data_type.append('����')

    data_make['����'] = data_type  # ������������
    # print(data_type)
    data_make['�ϰ�ʱ��'] = data_col_arrivetime  # �����ϰ�ʱ��
    data_make['�°�ʱ��'] = data_col_leavetime  # �����°�ʱ��
    data_make.�ϰ�ʱ�� = data_make.�ϰ�ʱ��.str.replace('OFF|���|����ְȱ��|����|ҽ����|�¼�|���|����|����|�����|ɥ��|�����', '', regex=True)#|���|����ְȱ��|����|ҽ����|�¼�|���|����|����|�����|ɥ��|�����
    data_make.�°�ʱ�� = data_make.�ϰ�ʱ��.str.replace('OFF|���|����ְȱ��|����|ҽ����|�¼�|���|����|����|�����|ɥ��|�����', '', regex=True)

    # ~ #��ȡ'����(YYYY-MM-DD)'--------����1
    # ~ data_time = list(map(lambda x: '2019-'+x, data_time))#�ù�ʽ��ÿ���ַ�ǰ��2019-

    data_yyyy = []
    for i in range(0, len(list(data.columns))):
        for data_t in list(data_time[2:]):
            data_yyyy.append(str(data_t)[:-9])  # ��ÿ��col��datatime�����
    data_make['����(YYYY-MM-DD)'] = data_yyyy  # ��ֵ��col����

    # ��openpyxl���ø�ʽ
    wb = Workbook()
    ws = wb.create_sheet('Sheet1', -1)
    for r in dataframe_to_rows(data_make, index=False, header=True):
        ws.append(r)
    cell_style(ws, len(data_make.index))
    wb.save(in_excel)
    return in_excel


file_name = '����8.5-8.11.xlsx'
wash_data(file_name)
