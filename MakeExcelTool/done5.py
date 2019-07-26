# -*-coding:GBK -*-
#  ����ǰ�α�
import numpy as np
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
pd.options.mode.chained_assignment = None


def cell_style_a5(ws, len_index, size_a):
	width_dict_a5 = {'A': 16, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16}
	font_a5 = Font(name='΢���ź�', size=8, bold=False)
	width_dict_a4 = {'A': 21.18, 'B': 21.18, 'C': 21.18, 'D': 21.18, 'E': 21.18, 'F': 21.18}
	font_a4 = Font(name='΢���ź�', size=11, bold=False)
	alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # wrap_text�Զ�����

	#  ����cell��ʽ
	if size_a == 'A5':
		font = font_a5
		width_dict = width_dict_a5
		height = 60
	else:
		font = font_a4
		width_dict = width_dict_a4
		height = 87.75
		
	for row in ws.iter_rows(min_row=1, max_row=len_index, min_col=0, max_col=6):
		for cell in row:
			cell.font = font
			cell.alignment = alignment
	#  �����п�
	for k, v in width_dict.items():
		ws.column_dimensions[k].width = v
	#  �����и�
	for i in range(len_index+1):
		ws.row_dimensions[i].height = height


#  ����ʱ��������һ����index
def time_sort_index(time_col):
	time_col = time_col.str.replace('4*[\u4e00-\u9fa5]|[:].*$', '')
	time_index = sorted(list(set(time_col)))
	return time_index


#  �����ַ���
def handle_str(teacher, kind_class, class_room):
	teacher = teacher.str.replace('[0-9]*$', '')  # ɾ����ʦ�������
	# ~ kind_class = kind_class.str.replace('^[\u4e00-\u9fa5][\u4e00-\u9fa5][\u4e00-\u9fa5]','')#ɾ����ѵ��������
	kind_class = kind_class.str.replace('�ּ��Ķ�', '�ּ��Ķ�˫ʦ')  # ɾ����ѵ��������
	class_room = class_room.str.replace('4*[\u4e00-\u9fa5]|[(].*?[)]|[��].*?[��]', '')  # ֻ�������Һ�
	return teacher, kind_class, class_room


def ceshi(data):
	writer = pd.ExcelWriter('ceshi.xlsx')
	data.to_excel(writer, index=True, header=True)
	writer.save()


def wash_data(filename):
	data = pd.read_excel(filename, sheet_name='Sheet0', usecols=[4, 6, 7, 9, 11, 14, 17, 18, 22])  # ��ȡ��
	
	timelist = time_sort_index(data['�Ͽ�ʱ��'])  # �Ͽ�ʱ��Ϊindex
	choose = data.iloc[1, 0]
	
	data_fudao = data['������ʦ'].str.replace('[0-9]*$', '')  # ɾ��������ʦ���ƺ������
	data_fudao2 = data_fudao.str.replace('.*[\u4e00-\u9fa5]', '����:', regex=True)  # ɾ�����ֺ�ֻʣ���֣�����ȫ���滻�ɸ���
	data['������ʦ'] = data_fudao2 + data_fudao  # ����+��ʦ����
	
	finish_excel = '���ȱ�(û��).xlsx'  # ��ȡ����Ĺ����ļ��� #finish_excel = data.loc[2,'��ѧ��']+'__'+'���ȱ�(û��).xlsx'
	in_excel = data.loc[2, '��ѧ��']+'__'+data.loc[2, 'ѧ��']+'__'+'������ǰ�α�.xlsx'  # ��ȡ�����ļ����������ݳ�ȥ
	
	data['��ʦ'], data['���'], data['����'] = handle_str(data['��ʦ'], data['���'], data['����'])  # ���������ַ�
	classroomlist = sorted(list(set(data['����'])))  # ���Һ�Ϊsheet_name
	classroomlist = [x for x in classroomlist if x != '']
	
	data2 = data['����']+data['�꼶']+data['ѧ��']+'\n'+data['���']+'\n'+data['��ʦ']+' '+data['������ʦ'] + '\n'+data['�Ͽ�ʱ��']

	writer = pd.ExcelWriter(finish_excel)
	num = []
	for i in range(1, len(timelist)+1):
		num.append(np.nan)	
	datamake_chunqiu = pd.DataFrame({'�ܶ�': num, '����': num, '����': num, '����': num, '����': num, '����': num}, index=timelist)
							
	datamake_hanshu = pd.DataFrame({'����': num, 'һ��': num, '����': num, '����': num, '����': num}, index=timelist)
	
	if choose == '������' or choose == '�＾��':
		data_make = datamake_chunqiu
	else:
		data_make = datamake_hanshu
	
	for class_list in classroomlist:
		data_class = data2[data2.str.contains(class_list)]
		data_class = data_class.replace(to_replace=r'^'+class_list, value='', regex=True)
		for col in list(data_make.columns):
			for ind in timelist:
				a = data_class[data_class.str.contains(col+'[\u4e00-\u9fa5][\u4e00-\u9fa5]'+ind+'[:].*$', regex=True)]
				b = list(a)
				if not b:
					data_make.loc[ind, col] = np.nan
				else:
					data_make.loc[ind, col] = b[0]
		data_make.to_excel(writer, sheet_name=class_list, index=True, header=True)
	writer.save()
	return classroomlist, finish_excel, in_excel
		
	
def wash_data2_hander1(data):
	data.dropna(axis=0, how='all', inplace=True)
	return data


def wash_data2(put_excel, in_excel, classroomlist, size_a):
	wb = Workbook()
	for cla in classroomlist:
		data = pd.read_excel(put_excel, sheet_name=cla, index_col=0)
		data = wash_data2_hander1(data)
		if len(data.index) > 5:
			ws = wb.create_sheet(cla+'(�����)', -1)
			for r in dataframe_to_rows(data, index=False, header=False):
				ws.append(r)
			cell_style_a5(ws, len(data.index), size_a)
		elif len(data.index) < 5:
			ws = wb.create_sheet(cla+'(������)', -1)
			for r in dataframe_to_rows(data, index=False, header=False):
				ws.append(r)
			cell_style_a5(ws, len(data.index), size_a)
		else:
			ws = wb.create_sheet(cla, -1)
			for r in dataframe_to_rows(data, index=False, header=False):
				ws.append(r)
			cell_style_a5(ws, len(data.index), size_a)
	wb.save(size_a + in_excel)
		

def final_fuc(filename, size_a):
	classroomlist, put_excel, in_excel = wash_data(filename)
	wash_data2(put_excel, in_excel, classroomlist, size_a)
	os.remove('���ȱ�(û��).xlsx')
	return in_excel

